import openpyxl
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, Button

root = tk.Tk()
global file_list
file_list = []
global file_names
file_names = []


def main():
    start_gui()

    # wb = openpyxl.load_workbook('Flyback 1_2018-07-03_0802.xlsx')
    # print(wb.sheetnames)
    # wb = modify_workbook(wb)
    # wb.save('test.xlsx')

'''
The code below handles the GUI
'''
def start_gui():
    style = ttk.Style(root)
    style.theme_use("clam")

    select_file = Button(root, text="Select files to convert", command=show_file_chooser)
    select_file.grid(row=1, padx=40, pady=2)

    convert = Button(root, text="Convert files", command=check_and_modify)
    convert.grid(row=2, padx=4, pady=2)

    root.mainloop()

def show_file_chooser():
    global file_list
    global file_names
    file_list = filedialog.askopenfilenames(parent=root, initialdir='/', initialfile='tmp', filetypes=[("XLSX", "*.xlsx"), ("All files", "*")])
    for file in file_list:
        name = file.split('/')[len(file.split('/')) - 1]
        file_names.append(name)
    return file_list

def check_and_modify():
    global file_list
    for file in file_list:
        wb = openpyxl.load_workbook(file)
        name = file.split('/')[len(file.split('/')) - 1]
        if is_workbook(wb):
            print(name + '\nConverting...\n')
            wb = modify_workbook(wb)
            wb.save(file.split('.xlsx')[0] + '_converted.xlsx')
        else:
            print(name + "\nWas not recognized\n")

    print('Complete!')

'''
The code below is for determining if a given wb is the correct type to convert
'''

'''
Ensures that wb has the structure of ['Collection', 'Images', 'POA Measurements'],
then it calls sheet specific checks.
'''
def is_workbook(wb):
    sheets = wb.sheetnames
    if not sheets[0] == 'Collection' or not sheets[1] == 'Images' or not sheets[2] == 'POA Measurements':
        return False
    elif not is_collection(wb['Collection']):
        return False
    elif not is_images(wb['Images']):
        return False
    elif not is_poa_measurements(wb['POA Measurements']):
        return False
    else:
        return True

def is_collection(ws):
    a1 = ws['A1'].value
    i1 = ws['I1'].value
    v1 = ws['V1'].value
    if not a1 == 'collectionId' or not i1 == 'Pole ID' or not v1 == 'Photo Measure.altitude':
        return False
    else:
        return True

def is_images(ws):
    c1 = ws['C1'].value
    k1 = ws['K1'].value
    r1 = ws['R1'].value
    if not c1 == 'type' or not k1 == 'compositeImageUrl'or not r1 == 'distance.display':
        return False
    else:
        return True

def is_poa_measurements(ws):
    a1 = ws['A1'].value
    f1 = ws['F1'].value
    h1 = ws['H1'].value
    if not a1 == 'collectionId' or not f1 == 'POA Height' or not h1 == 'Comments':
        return False
    else:
        return True



'''
The code below is for converting an unconverted .xlsx file with the structure
['Collection', 'Images', 'POA Measurements']
'''

'''
Wrapper method for calling sheet specific methods

Input : wb = unconverted .xlsx
Output: wb = converted .xlsx
'''
def modify_workbook(wb):
     wb = modify_collection(wb)

     wb = modify_images(wb)

     wb = modify_poa(wb)
     return wb

'''
Converts a workbook with an unconverted 'Collection' sheet

Input : wb = workbook with an unconverted 'Collection' sheet
Output: wb = workbook with a converted 'Collection' sheet
'''
def modify_collection(wb):
    ws = wb['Collection']
    height = len(tuple(ws.columns)[0])

    # Modify column names
    ws['A1'].value = 'ID'
    ws['I1'].value = 'PoleID'
    ws['J1'].value = 'Height'
    ws['L1'].value = 'Latitude'
    ws['N1'].value = 'Longitude'
    ws['Y1'].value = 'Material'
    ws['Z1'].value = 'Comment'

    ws = shift_column('I', 'B', height, ws)
    ws = shift_column('J', 'C', height, ws)
    ws = shift_column('K', 'D', height, ws)
    ws = shift_column('L', 'E', height, ws)
    ws = shift_column('N', 'F', height, ws)
    ws = shift_column('Y', 'G', height, ws)
    ws = shift_column('Z', 'H', height, ws)
    ws = shift_column('AA', 'I', height, ws)

    for row in ws['J1:AA{0}'.format(height)]:
        for cell in row:
            cell.value = ''

    ws['J1'].value = 'PhotoFile'

    return wb

'''
Converts a workbook with an unconverted 'Images' sheet

Input : wb = workbook with an unconverted 'Images' sheet
Output: wb = workbook with a converted 'Images' sheet
'''
def modify_images(wb):
    ws = wb['Images']
    height = len(tuple(ws.columns)[0])

    ws = shift_column('D', 'B', height, ws)
    ws = shift_column('I', 'C', height, ws)

    for row in ws['D1:R{0}'.format(height)]:
        for cell in row:
            cell.value = ''
    
    return wb

'''
Converts a workbook with an unconverted 'POA Measurements' sheet

Input : wb = workbook with an unconverted 'POA Measurements' sheet
Output: wb = workbook with a converted 'POA Measurements' sheet
'''
def modify_poa(wb):
    ws = wb['POA Measurements']
    height = len(tuple(ws.columns)[0])

    # Insert VLOOKUP for Latitude
    for row in ws['I1:I{0}'.format(height)]:
        for cell in row:
            cell.value = "=VLOOKUP(A{0}, Collection!A:G,5,FALSE)".format(cell.row)

    # Insert VLOOKUP for Longitude
    for row in ws['J1:J{0}'.format(height)]:
        for cell in row:
            cell.value = "=VLOOKUP(A{0}, Collection!A:G,6,FALSE)".format(cell.row)

    # Modify column names
    ws['A1'].value = 'LinkedID'
    ws['B1'].value = 'PoleID'
    ws['D1'].value = 'ID'
    ws['E1'].value = 'Type'
    ws['F1'].value = 'Height'
    ws['H1'].value = 'Comment'
    ws['I1'].value = 'Latitude'
    ws['J1'].value = 'Longitude'
    ws['K1'].value = 'POAID'
    
    return wb

'''
Moves all cells in column a into corresponding cells in column b

Input : a = single character string of starting column
        b = single character string of ending column
        height = total number of rows in the given sheet
        ws = the sheet to modify
Output: ws = the modified sheet
'''
def shift_column(a, b, height, ws):
    for row in ws['{0}1:{0}{1}'.format(b, height)]:
        for cell in row:
            cell.value = ws['{0}{1}'.format(a, cell.row)].value
    return ws

main()